"""
Excel → 일별 JSON 데이터 변환기 (Q1)

Excel 파일 형식:
  시트명 = 날짜 (예: 20260512) 또는 'template'
  각 시트에 카테고리별 테이블 포함

사용법:
  python excel_to_data.py market_data.xlsx
  python excel_to_data.py market_data.xlsx --date 20260512
  python excel_to_data.py --template            ← 입력용 엑셀 템플릿 생성

출력:
  data/daily/YYYYMMDD/*.json  (8개 파일)
"""

import json, sys
from pathlib import Path

BASE = Path(__file__).parent

# ── Excel 의존성 확인 ─────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("[ERROR] openpyxl이 설치되지 않았습니다.")
    print("  실행: pip install openpyxl")
    sys.exit(1)


# ── 템플릿 엑셀 생성 ─────────────────────────────────────────────────

TEMPLATE_SHEETS = {
    "국내금리": {
        "headers": ["항목명", "금리(%)"],
        "rows": [
            ["CD 91D", ""], ["통안채 1Y", ""], ["통안채 2Y", ""],
            ["국고채 3Y", ""], ["국고채 5Y", ""], ["국고채 10Y", ""],
            ["은행채 3M", ""], ["은행채 1Y", ""], ["은행채 2Y", ""],
            ["은행채 3Y", ""], ["은행채 5Y", ""],
            ["회사채 1Y (AA)", ""], ["회사채 3Y (AA)", ""], ["기타금융채 2Y (AA-)", ""],
        ]
    },
    "국내주식환율": {
        "headers": ["항목명", "값"],
        "rows": [["KOSPI",""],["KOSPI 200",""],["KOSDAQ",""],["USDKRW",""]],
    },
    "투자자동향": {
        "headers": ["시장", "외국인(억원)", "기관(억원)", "개인(억원)"],
        "rows": [
            ["KOSPI","","",""],["KOSDAQ","","",""],
            ["KOSPI 200 선물","","",""],["국채 3년 선물","","",""],["국채 10년 선물","","",""],
        ],
    },
    "해외금리": {
        "headers": ["항목명", "금리(%)"],
        "rows": [["미국 2Y",""],["미국 10Y",""],["미국 30Y",""],["독일 10Y",""],["일본 10Y",""]],
    },
    "해외주식환율암호화폐": {
        "headers": ["항목명", "값", "구분"],
        "rows": [
            ["DOW","","주식"],["S&P 500","","주식"],["NASDAQ","","주식"],
            ["일본 니케이 225","","주식"],["홍콩 항셍 H","","주식"],["독일 DAX","","주식"],
            ["달러인덱스","","FX"],["USDJPY","","FX"],["EURUSD","","FX"],
            ["비트코인(USD)","","암호화폐"],["이더리움(USD)","","암호화폐"],
        ],
    },
    "상품": {
        "headers": ["항목명", "값"],
        "rows": [["WTI",""],["BRENT",""],["금",""],["은",""],["PHIL 반도체지수",""],["구리",""]],
    },
    "코멘터리": {
        "headers": ["구분", "내용"],
        "rows": [["국내",""],["해외",""]],
    },
}

def create_template(out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="003087")

    for sheet_name, cfg in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        # 날짜 입력란
        ws["A1"] = "보고서날짜(YYYYMMDD)"
        ws["B1"] = ""
        ws["A1"].font = Font(bold=True)

        # 헤더
        for col, h in enumerate(cfg["headers"], 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        # 데이터 행
        for r, row in enumerate(cfg["rows"], 4):
            for col, val in enumerate(row, 1):
                ws.cell(row=r, column=col, value=val)

        # 열 너비 조정
        ws.column_dimensions["A"].width = 22
        for col in "BCDE":
            ws.column_dimensions[col].width = 16

    wb.save(out_path)
    print(f"[OK] 템플릿 생성: {out_path}")
    print("  각 시트의 B1 셀에 보고서 날짜(YYYYMMDD)를 입력하고 데이터를 채운 후 다시 실행하세요.")


# ── Excel 파싱 ────────────────────────────────────────────────────────

def safe_float(v):
    try: return float(v) if v not in (None, "", "-") else None
    except (ValueError, TypeError): return None

def safe_int(v):
    try: return int(float(v)) if v not in (None, "", "-") else None
    except (ValueError, TypeError): return None

def parse_excel(wb, report_date: str) -> dict:
    """
    엑셀에서 데이터를 읽어 daily JSON 구조로 변환.
    B1 셀에 데이터 날짜 입력, 없으면 report_date 사용.
    """
    # 이전 날 데이터가 없으므로 prev_day 변화는 None으로 설정
    # 사용자가 이후 collect_daily.py로 채울 수 있음

    result = {}
    sheet_map = {s.title: s for s in wb.worksheets}

    def get_ws(name):
        for k, v in sheet_map.items():
            if k.strip() == name.strip():
                return v
        return None

    # ── 공통: 날짜 추출 ──────────────────────────────────────────────
    first_ws = list(sheet_map.values())[0] if sheet_map else None
    data_date = None
    if first_ws:
        raw = first_ws["B1"].value
        if raw:
            s = str(raw).replace("-","")
            if len(s) == 8 and s.isdigit():
                data_date = f"{s[:4]}-{s[4:6]}-{s[6:]}"
    if not data_date:
        data_date = f"{report_date[:4]}-{report_date[4:6]}-{report_date[6:]}"

    def no_change(val_str):
        return {"prev_day_bp": None, "ytd_bp": None,
                "prev_day_str": "—", "ytd_str": "—",
                "prev_day_pct": None, "ytd_pct": None}

    # ── 국내 금리 ────────────────────────────────────────────────────
    ws = get_ws("국내금리")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, val = row[0], row[1]
            if not name: continue
            v = safe_float(val)
            items.append({"name": str(name), "value": v, **no_change(v)})
        result["domestic_rates"] = {
            "data_date": data_date, "items": items, "source": "Excel 입력"
        }

    # ── 국내 주식/환율 ───────────────────────────────────────────────
    ws = get_ws("국내주식환율")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, val = row[0], row[1]
            if not name: continue
            v = safe_float(val)
            items.append({"name": str(name), "value": v, **no_change(v)})
        result["domestic_markets"] = {
            "data_date": data_date, "items": items, "source": "Excel 입력"
        }

    # ── 투자자 동향 ──────────────────────────────────────────────────
    ws = get_ws("투자자동향")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            market, f, i, r = row[0], row[1], row[2], row[3]
            if not market: continue
            items.append({
                "market": str(market),
                "foreign":     safe_int(f),
                "institution": safe_int(i),
                "individual":  safe_int(r),
            })
        result["investor_flow"] = {
            "data_date": data_date, "unit": "억원", "items": items, "source": "Excel 입력"
        }

    # ── 해외 금리 ────────────────────────────────────────────────────
    ws = get_ws("해외금리")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, val = row[0], row[1]
            if not name: continue
            v = safe_float(val)
            items.append({"name": str(name), "value": v, **no_change(v)})
        result["overseas_rates"] = {
            "data_date": data_date, "items": items, "source": "Excel 입력"
        }

    # ── 해외 주식/환율/암호화폐 ──────────────────────────────────────
    ws = get_ws("해외주식환율암호화폐")
    if ws:
        stocks, fx, crypto = [], [], []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, val, cat = row[0], row[1], row[2] if len(row) > 2 else None
            if not name: continue
            v = safe_float(val)
            item = {"name": str(name), "value": v, **no_change(v)}
            cat = str(cat or "").strip()
            if cat == "FX":       fx.append(item)
            elif cat == "암호화폐": crypto.append(item)
            else:                 stocks.append(item)
        result["overseas_markets"] = {
            "data_date": data_date, "stocks": stocks, "fx": fx,
            "crypto": crypto, "source": "Excel 입력"
        }

    # ── 상품 ─────────────────────────────────────────────────────────
    ws = get_ws("상품")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, val = row[0], row[1]
            if not name: continue
            v = safe_float(val)
            items.append({"name": str(name), "value": v, **no_change(v)})
        result["commodities"] = {
            "data_date": data_date, "items": items, "source": "Excel 입력"
        }

    # ── 코멘터리 ─────────────────────────────────────────────────────
    ws = get_ws("코멘터리")
    if ws:
        items = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            tag, text = row[0], row[1]
            if not tag: continue
            items.append({"tag": str(tag), "text": str(text or "")})
        result["commentary"] = {"data_date": data_date, "items": items}

    # ── 크레딧 스프레드 (자동 계산) ──────────────────────────────────
    dom_r = result.get("domestic_rates", {})
    items = dom_r.get("items", [])
    gov2y_item  = next((x for x in items if "통안" in x["name"] and "2Y" in x["name"]), None)
    corp3y_item = next((x for x in items if "회사채" in x["name"] and "3Y" in x["name"]), None)
    if gov2y_item and corp3y_item:
        g = gov2y_item["value"] or 0
        c = corp3y_item["value"] or 0
        result["credit_spread"] = {
            "data_date": data_date,
            "chart_labels": ["—"],
            "gov_2y":     [g],
            "corp_aa_2y": [c],
            "spread":     [round(c - g, 3)],
            "source": "Excel 입력"
        }

    return result, data_date


def save_daily_files(parsed: dict, report_date: str):
    out = BASE / "data" / "daily" / report_date
    out.mkdir(parents=True, exist_ok=True)

    file_map = {
        "domestic_rates":   "domestic_rates.json",
        "domestic_markets": "domestic_markets.json",
        "investor_flow":    "investor_flow.json",
        "overseas_rates":   "overseas_rates.json",
        "overseas_markets": "overseas_markets.json",
        "commodities":      "commodities.json",
        "commentary":       "commentary.json",
        "credit_spread":    "credit_spread.json",
    }
    written = []
    for key, fname in file_map.items():
        if key in parsed:
            path = out / fname
            path.write_text(json.dumps(parsed[key], ensure_ascii=False, indent=2), encoding="utf-8")
            written.append(fname)

    return written


# ── 메인 ─────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if "--template" in args:
        out = BASE / "market_data_template.xlsx"
        create_template(out)
        return

    if not args:
        print("사용법:")
        print("  python excel_to_data.py market_data.xlsx")
        print("  python excel_to_data.py market_data.xlsx --date 20260512")
        print("  python excel_to_data.py --template")
        return

    xlsx_path = Path(args[0])
    if not xlsx_path.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {xlsx_path}")
        return

    # 보고서 날짜 결정
    date_arg = None
    for i, a in enumerate(args):
        if a == "--date" and i + 1 < len(args):
            date_arg = args[i + 1].replace("-","")

    if not date_arg:
        # 파일명에서 날짜 추출 시도 (예: data_20260512.xlsx)
        stem = xlsx_path.stem
        digits = "".join(c for c in stem if c.isdigit())
        date_arg = digits[:8] if len(digits) >= 8 else None

    if not date_arg:
        print("[ERROR] 보고서 날짜를 알 수 없습니다. --date YYYYMMDD 옵션을 사용하세요.")
        return

    print(f"[엑셀 로드] {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    parsed, data_date = parse_excel(wb, date_arg)
    written = save_daily_files(parsed, date_arg)

    print(f"[OK] 보고서 날짜: {date_arg}, 데이터 날짜: {data_date}")
    print(f"[OK] 생성된 파일: {len(written)}개")
    for f in written:
        print(f"     data/daily/{date_arg}/{f}")

    print("\n다음 명령으로 대시보드를 재생성하세요:")
    print(f"  python generate_daily.py {date_arg}")


if __name__ == "__main__":
    main()
